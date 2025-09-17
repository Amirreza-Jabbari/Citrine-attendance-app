# src/citrine_attendance/services/export_service.py
import csv
import logging
from pathlib import Path
from typing import List, Dict, Any
import datetime
import jdatetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from ..config import config
from ..utils.time_utils import minutes_to_hhmm
from ..utils.resources import get_resource_path
from ..locale import _

class ExportServiceError(Exception):
    pass

class ExportService:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self._register_persian_font()

    def _register_persian_font(self):
        """Registers the Vazir font for PDF exports if it exists."""
        try:
            font_path = get_resource_path('fonts/Vazir-Regular.ttf')
            pdfmetrics.registerFont(TTFont('Vazir', font_path))
        except Exception as e:
            self.logger.warning(f"Could not register Persian font: {e}")

    def export_data(self, export_format: str, data: List[Dict[str, Any]], path: Path, title: str = "Attendance Report"):
        """Dispatches the export request to the correct method based on the format."""
        try:
            translated_title = _("attendance_report_title")
            if export_format == 'csv':
                return self.export_to_csv(data, path)
            elif export_format == 'xlsx':
                return self.export_to_xlsx(data, path, title=translated_title)
            elif export_format == 'pdf':
                return self.export_to_pdf(data, path, title=translated_title)
            else:
                raise ExportServiceError(f"Unsupported export format: {export_format}")
        except Exception as e:
            self.logger.error(f"Export dispatcher failed for format {export_format}: {e}", exc_info=True)
            raise

    def _process_data_for_export(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Process data for export, formatting dates and converting all minute fields to HH:MM.
        """
        processed_data = []
        date_format_pref = config.settings.get("date_format", "both")

        # HEROIC FIX: This map now directly uses the translated keys.
        minute_keys_map = {
            _("Tardiness (min)"): _("Tardiness (H:M)"),
            _("Early Departure (min)"): _("Early Departure (H:M)"),
            _("Main Work (min)"): _("Main Work (H:M)"),
            _("Overtime (min)"): _("Overtime (H:M)"),
            _("Launch Time (min)"): _("Launch Time (H:M)"),
            _("Total Duration (min)"): _("Total Duration (H:M)"),
            _("Leave (min)"): _("Leave (H:M)"),
            _("Used Leave This Month (min)"): _("Used Leave This Month (H:M)"),
            _("Remaining Leave This Month (min)"): _("Remaining Leave This Month (H:M)")
        }

        for row in data:
            processed_row = row.copy()
            
            # Handle date formatting
            date_key = _("Date")
            greg_date_str = processed_row.get(date_key)
            if isinstance(greg_date_str, str):
                try:
                    greg_date = datetime.date.fromisoformat(greg_date_str)
                    if date_format_pref == 'jalali':
                        processed_row[date_key] = jdatetime.date.fromgregorian(date=greg_date).strftime("%Y/%m/%d")
                    elif date_format_pref == 'gregorian':
                        processed_row[date_key] = greg_date.isoformat()
                    else: 
                        j_date_str = jdatetime.date.fromgregorian(date=greg_date).strftime("%Y/%m/%d")
                        g_date_str = greg_date.isoformat()
                        processed_row[date_key] = f"{j_date_str} | {g_date_str}"
                except (ValueError, TypeError):
                    pass # Keep original string if parsing fails
            
            # Handle time formatting
            for key in [_("Time In"), _("Time Out")]:
                time_val = processed_row.get(key)
                if isinstance(time_val, datetime.time):
                    processed_row[key] = time_val.strftime("%H:%M")

            # Handle minute to H:M conversion
            for min_key, hm_key in minute_keys_map.items():
                if min_key in processed_row:
                    minutes_val = processed_row.pop(min_key)
                    processed_row[hm_key] = minutes_to_hhmm(minutes_val)
            
            processed_data.append(processed_row)
        return processed_data

    def export_to_csv(self, data: List[Dict[str, Any]], filename: Path, delimiter: str = ',') -> Path:
        """Export data to a CSV file."""
        try:
            if not data:
                raise ExportServiceError("No data provided for CSV export.")
            
            processed_data = self._process_data_for_export(data)
            if not processed_data:
                raise ExportServiceError("No data available after processing.")

            fieldnames = processed_data[0].keys()
            
            with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=delimiter)
                writer.writeheader()
                writer.writerows(processed_data)

            self.logger.info(f"Data exported to CSV: {filename}")
            return filename
        except Exception as e:
            self.logger.error(f"Error exporting to CSV: {e}", exc_info=True)
            raise ExportServiceError(f"Failed to export to CSV: {e}") from e

    def export_to_xlsx(self, data: List[Dict[str, Any]], filename: Path, title: str) -> Path:
        """Export data to an Excel (XLSX) file with formatting."""
        try:
            if not data:
                raise ExportServiceError("No data provided for XLSX export.")
            
            processed_data = self._process_data_for_export(data)
            if not processed_data:
                raise ExportServiceError("No data available after processing.")

            wb = Workbook()
            ws = wb.active
            ws.title = title

            # Set sheet direction for RTL languages
            if config.settings.get("language", "en") == "fa":
                ws.sheet_view.rightToLeft = True

            headers = list(processed_data[0].keys())
            ws.append(headers)
            
            for row_dict in processed_data:
                row_data = [row_dict.get(h, "") for h in headers]
                ws.append(row_data)

            header_font = Font(bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, column_title in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                
                max_length = len(str(column_title))
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=col_num).value
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = max_length + 4
                ws.column_dimensions[get_column_letter(col_num)].width = min(adjusted_width, 40)
                
                # Heuristic to find time-related columns for centering
                if any(sub in str(column_title).lower() for sub in ["(h:m)", "time", "(ساعت)", "زمان"]):
                    for row in range(1, ws.max_row + 1):
                        ws.cell(row=row, column=col_num).alignment = center_alignment

            wb.save(filename)
            self.logger.info(f"Data exported to XLSX: {filename}")
            return filename
        except Exception as e:
            self.logger.error(f"Error exporting to XLSX: {e}", exc_info=True)
            raise ExportServiceError(f"Failed to export to XLSX: {e}") from e

    def export_to_pdf(self, data: List[Dict[str, Any]], filename: Path, title: str) -> Path:
        """Export data to a PDF file, handling RTL for Persian language."""
        try:
            if not data:
                raise ExportServiceError("No data provided for PDF export.")

            doc = SimpleDocTemplate(str(filename), pagesize=landscape(A4))
            elements = []
            
            is_persian = config.settings.get("language", "en") == "fa"
            font_name = 'Vazir' if is_persian else 'Helvetica'
            bold_font_name = 'Vazir' if is_persian else 'Helvetica-Bold'

            styles = getSampleStyleSheet()
            title_align = 2 if is_persian else 1 # 2=RIGHT, 1=CENTER
            styles.add(ParagraphStyle(name='Title_Custom', parent=styles['Title'], fontName=font_name, alignment=title_align))
            styles.add(ParagraphStyle(name='Body_Custom', parent=styles['Normal'], fontName=font_name, alignment=2))

            elements.append(Paragraph(title, styles['Title_Custom']))
            elements.append(Spacer(1, 0.2*inch))
            
            processed_data = self._process_data_for_export(data)
            if not processed_data:
                raise ExportServiceError("No data available after processing.")
                
            headers = list(processed_data[0].keys())
            table_data = [[str(row.get(h, "")) for h in headers] for row in processed_data]

            if is_persian:
                # Reverse for RTL display
                headers.reverse()
                table_data = [row[::-1] for row in table_data]

            # Re-insert headers at the top
            table_data.insert(0, headers)
            
            table = Table(table_data, repeatRows=1)
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), bold_font_name),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ])
            table.setStyle(style)

            # Alternating row colors
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    table.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), colors.lightgrey)]))
            
            elements.append(table)
            doc.build(elements)

            self.logger.info(f"Data exported to PDF: {filename}")
            return filename
        except Exception as e:
            self.logger.error(f"Error exporting to PDF: {e}", exc_info=True)
            raise ExportServiceError(f"Failed to export to PDF: {e}") from e

# Global instance
export_service = ExportService()